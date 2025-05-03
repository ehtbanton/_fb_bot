import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os.path
import calendar

def update_sheet_range():
    """
    Updates the Excel sheet with a rolling calendar for the next year.
    Voice parts in column B, dates in row 1 starting from column C.
    """
    # File path
    file_path = "dep_calendar.xlsx"
    
    # Voice parts
    voice_parts = ["DA1", "DA2", "CA1", "CA2", "DT1", "DT2", "CT1", "CT2", 
                   "DB1", "DB2", "DB3", "CB1", "CB2", "CB3"]
    
    # Create or load workbook
    if os.path.isfile(file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Deputy Bookings"
    
    # Set column widths
    for col_idx in range(1, 10):
        column_letter = get_column_letter(col_idx)
        sheet.column_dimensions[column_letter].width = 15
    
    # Get today's date
    today = datetime.now()
    
    # Define the start date (this week's Monday)
    start_date = today - timedelta(days=today.weekday())
    
    # Clear existing data
    sheet.delete_rows(1, sheet.max_row)
    
    # Row 1: Date headers starting from column C
    row_num = 1
    current_date = start_date
    end_date = start_date + timedelta(days=365)
    
    while current_date < end_date:
        # Add date headers (Mon, Tue, etc.)
        for day_idx in range(7):
            date = current_date + timedelta(days=day_idx)
            day_name = calendar.day_name[date.weekday()][:3]  # Mon, Tue, etc.
            
            # Format the day (1st, 2nd, 3rd, etc.)
            day_num = date.day
            if 4 <= day_num <= 20 or 24 <= day_num <= 30:
                suffix = "th"
            else:
                suffix = {1: "st", 2: "nd", 3: "rd"}.get(day_num % 10, "th")
            
            month_name = calendar.month_name[date.month][:3]  # Jan, Feb, etc.
            date_header = f"{day_name} {day_num}{suffix} {month_name}"
            
            cell = sheet.cell(row=row_num, column=day_idx + 3)  # Start from column C (index 3)
            cell.value = date_header
            cell.alignment = Alignment(horizontal='center')
        
        # Add voice parts in column B and color cells green
        for i, part in enumerate(voice_parts):
            voice_row = row_num + 1 + i
            sheet.cell(row=voice_row, column=2).value = part  # Column B
            
            # Color cells green for all days
            for day_idx in range(7):
                col = day_idx + 3  # Start from column C (index 3)
                cell = sheet.cell(row=voice_row, column=col)
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        # Move to the next week
        current_date += timedelta(days=7)
        row_num += len(voice_parts) + 2  # Add empty row after each week
    
    # Save the workbook
    wb.save(file_path)
    print(f"Calendar updated in {file_path}")
    return wb

def update_cell(date_str, voice_part, text="", colour="green"):
    """
    Updates a specific cell on the calendar.
    
    Args:
        date_str (str): Date in format "DD-MM-YYYY" (e.g., "05-05-2025")
        voice_part (str): Voice part (e.g., "CB2", "DT1")
        text (str): Text to display in the cell
        colour (str): Cell color - "green", "yellow", "black" or any valid hex color
    """
    # Load the workbook
    file_path = "dep_calendar.xlsx"
    if not os.path.isfile(file_path):
        # Create the sheet first if it doesn't exist
        wb = update_sheet_range()
    else:
        wb = openpyxl.load_workbook(file_path)
        
    sheet = wb.active
    
    # Parse the input date
    day, month, year = map(int, date_str.split('-'))
    target_date = datetime(year, month, day)
    target_weekday = target_date.weekday()  # 0=Monday, 6=Sunday
    
    # Voice parts list
    voice_parts = ["DA1", "DA2", "CA1", "CA2", "DT1", "DT2", "CT1", "CT2", 
                   "DB1", "DB2", "DB3", "CB1", "CB2", "CB3"]
    
    # Find the week containing our target date
    found = False
    week_start_row = 1
    
    while week_start_row <= sheet.max_row:
        # Check if this row contains date headers
        header_cell = sheet.cell(row=week_start_row, column=3)  # Column C
        if not header_cell.value or "Mon" not in str(header_cell.value):
            week_start_row += 1
            continue
            
        # Parse the Monday date from the header
        monday_header = str(header_cell.value)
        try:
            parts = monday_header.split()
            day_num = int(''.join(filter(str.isdigit, parts[1])))
            month_str = parts[2]
            month_map = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
                         "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}
            month_num = month_map.get(month_str, 1)
            
            # Assume current year unless month is earlier than current month
            current_month = datetime.now().month
            year = datetime.now().year
            if month_num < current_month:
                year += 1
                
            monday_date = datetime(year, month_num, day_num)
            
            # Check if our target date is in this week
            days_diff = (target_date - monday_date).days
            if 0 <= days_diff < 7:
                found = True
                break
        except (ValueError, IndexError):
            pass
            
        # Move to the next week block
        week_start_row += len(voice_parts) + 2
    
    if not found:
        print(f"Week containing date {date_str} not found in calendar")
        return
    
    # Find the column for this date (C=Monday, D=Tuesday, etc.)
    target_col = target_weekday + 3  # +3 because we start from column C
    
    # Find the row for this voice part
    try:
        voice_idx = voice_parts.index(voice_part)
        target_row = week_start_row + 1 + voice_idx
    except ValueError:
        print(f"Voice part {voice_part} not found")
        return
    
    # Update the cell
    cell = sheet.cell(row=target_row, column=target_col)
    if text:
        cell.value = text
    
    # Set the color
    color_map = {
        "green": "00FF00",
        "yellow": "FFFF00",
        "black": "000000"
    }
    
    # Get the hex color code
    if colour in color_map:
        hex_color = color_map[colour]
    else:
        # Assume it's already a hex code
        hex_color = colour.lstrip('#')
    
    # Apply the fill
    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    
    # Save the workbook
    wb.save(file_path)
    print(f"Updated cell for {voice_part} on {date_str} to {colour} with text '{text}'")

# Example usage:
# update_sheet_range()  # Update/create the calendar
# update_cell("05-05-2025", "CB1", "Deputy", "yellow")  # Update a specific cell